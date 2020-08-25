import openpyxl, os
os.chdir("F://")
#print(os.listdir())
wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet3')
print(sheet)
print(type(sheet))
print(sheet.title)
anotherSheet = wb.active
#anotherSheet = wb.get_active_sheet() changed to wb.active
print(anotherSheet)
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print('Row' + str(c.row) + ', Column ' + str(c.column) +' is ' + str(c.value))
print('Cell' + c.coordinate + ' is ' +c.value)
print(sheet['C1'].value)

#specifying column and row using integers
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1,8,2):
    print(i, sheet.cell(row=i, column=2).value)
print(sheet.max_row)
print(sheet.max_column)

#Converting between Column Letter and Numbers
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(920))
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))

#Getting rows and columns from the sheets
print(tuple(sheet['A1' :'C3']))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('---END OF ROW ---')


#Access the value of cells in a particular row or columns
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active 
print(type(sheet.columns))
list(sheet.columns)[1]
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)







